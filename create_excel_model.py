#!/usr/bin/env python3
"""
MYOR Comparable Companies Analysis - Excel Model Generator
Investment Banking Best Practices Applied

Run: python create_excel_model.py
Output: MYOR_Comps_Model.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
import numpy as np

# =============================================================================
# STYLING CONSTANTS (IB Best Practices)
# =============================================================================

# Colors
NAVY = "1F4E79"
LIGHT_BLUE = "D6E3F8"
DARK_GRAY = "404040"
LIGHT_GRAY = "F2F2F2"
GREEN = "C6EFCE"
RED = "FFC7CE"
YELLOW = "FFEB9C"
WHITE = "FFFFFF"

# Fonts
HEADER_FONT = Font(name='Arial', size=10, bold=True, color=WHITE)
SUBHEADER_FONT = Font(name='Arial', size=9, bold=True, color=NAVY)
DATA_FONT = Font(name='Arial', size=9, color=DARK_GRAY)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color=NAVY)
HIGHLIGHT_FONT = Font(name='Arial', size=9, bold=True, color=NAVY)

# Fills
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type='solid')
ALTERNATE_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
HIGHLIGHT_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type='solid')
TARGET_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type='solid')

# Border
THIN_BORDER = Border(
    left=Side(style='thin', color=DARK_GRAY),
    right=Side(style='thin', color=DARK_GRAY),
    top=Side(style='thin', color=DARK_GRAY),
    bottom=Side(style='thin', color=DARK_GRAY)
)

# Alignment
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

# =============================================================================
# DATA
# =============================================================================

# Company data
companies = {
    'Company': ['Mayora Indah', 'Indofood CBP', 'Ultrajaya Milk', 'Garudafood',
                'Cisarua Mountain Dairy', 'Nippon Indosari', 'FKS Food'],
    'Ticker': ['MYOR IJ', 'ICBP IJ', 'ULTJ IJ', 'GOOD IJ', 'CMRY IJ', 'ROTI IJ', 'AISA IJ'],
    'Weight': ['Target', '25%', '20%', '19%', '16%', '11%', '7%'],
    'Price_IDR': [2120, 11000, 1425, 362, 6200, 800, 140],
    'Shares_B': [22.36, 11.66, 10.33, 38.12, 8.06, 6.19, 7.86],
    'Market_Cap_T': [47.4, 128.3, 14.73, 13.8, 50.0, 4.95, 1.1],
    'Net_Debt_T': [20.9, 39.2, -3.53, 1.4, -1.5, 1.15, 0.2],
    'EV_T': [68.3, 167.5, 11.2, 15.2, 48.5, 6.1, 1.3],
    '52Wk_High': [3010, 12500, 1800, 420, 6800, 1050, 165],
    '52Wk_Low': [1820, 9200, 1200, 280, 4800, 720, 95],
    '52Wk_Return': [-0.189, 0.082, -0.124, 0.225, 0.158, -0.180, 0.420],
}

financials = {
    'Company': ['Mayora Indah', 'Indofood CBP', 'Ultrajaya Milk', 'Garudafood',
                'Cisarua Mountain Dairy', 'Nippon Indosari', 'FKS Food'],
    'Revenue_T': [36.07, 72.6, 8.87, 12.24, 9.03, 3.93, 1.92],
    'Gross_Profit_T': [8.30, 26.93, 2.53, 3.49, 4.09, 2.12, 0.71],
    'EBITDA_T': [4.00, 16.70, 2.06, 1.41, 2.32, 0.72, 0.16],
    'Net_Income_T': [3.00, 7.10, 1.14, 0.62, 1.52, 0.36, 0.07],
    'Gross_Margin': [0.230, 0.371, 0.285, 0.285, 0.453, 0.538, 0.372],
    'EBITDA_Margin': [0.111, 0.230, 0.232, 0.115, 0.257, 0.183, 0.081],
    'Net_Margin': [0.083, 0.098, 0.129, 0.051, 0.168, 0.092, 0.036],
    'ROE': [0.153, 0.162, 0.164, 0.089, 0.261, 0.092, 0.065],
    'ROIC': [0.092, 0.124, 0.107, 0.062, 0.185, 0.078, 0.056],
}

multiples = {
    'Company': ['Mayora Indah', 'Indofood CBP', 'Ultrajaya Milk', 'Garudafood',
                'Cisarua Mountain Dairy', 'Nippon Indosari', 'FKS Food'],
    'EV_Revenue': [1.89, 2.31, 1.26, 1.24, 5.37, 1.55, 0.68],
    'EV_EBITDA': [17.08, 10.03, 5.44, 10.78, 20.91, 8.47, 8.13],
    'PE_TTM': [15.80, 14.83, 13.56, 21.43, 26.73, 13.51, 13.94],
    'PE_FWD': [14.60, 13.20, 12.80, 18.50, 22.80, 11.20, None],
    'PB': [2.90, 2.40, 1.77, 1.89, 5.44, 1.65, 1.00],
    'Div_Yield': [0.025, 0.028, 0.035, 0.022, 0.028, 0.082, 0.015],
}

growth = {
    'Company': ['Mayora Indah', 'Indofood CBP', 'Ultrajaya Milk', 'Garudafood',
                'Cisarua Mountain Dairy', 'Nippon Indosari', 'FKS Food'],
    'Rev_Growth_YoY': [0.062, 0.070, 0.069, 0.160, 0.161, 0.029, 0.127],
    'Rev_CAGR_3Y': [0.082, 0.100, 0.055, 0.098, 0.182, 0.042, 0.085],
    'EBITDA_Growth_YoY': [0.085, 0.130, 0.052, 0.185, 0.224, 0.058, 0.450],
    'EPS_Growth_YoY': [-0.060, 0.013, -0.028, 0.220, 0.223, 0.088, 2.700],
    'Fwd_Rev_Growth': [0.096, 0.080, 0.065, 0.120, 0.150, 0.050, 0.080],
    'Fwd_EPS_Growth': [0.191, 0.122, 0.059, 0.158, 0.171, 0.206, None],
}

leverage = {
    'Company': ['Mayora Indah', 'Indofood CBP', 'Ultrajaya Milk', 'Garudafood',
                'Cisarua Mountain Dairy', 'Nippon Indosari', 'FKS Food'],
    'Total_Debt_T': [9.30, 39.20, 0.28, 1.40, 0.75, 1.15, 0.20],
    'Net_Debt_EBITDA': [5.23, 2.35, -1.71, 0.99, -0.65, 1.60, 1.25],
    'Debt_Equity': [0.65, 0.98, 0.03, 0.19, 0.08, 0.38, 0.18],
    'Interest_Coverage': [4.71, 13.92, 42.00, 7.78, 115.00, 7.59, 6.20],
    'Current_Ratio': [2.50, 1.81, 4.22, 1.60, 2.80, 1.20, 1.40],
}

# =============================================================================
# CREATE WORKBOOK
# =============================================================================

def create_excel_model():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # ==========================================================================
    # SHEET 1: SUMMARY
    # ==========================================================================
    ws_summary = wb.create_sheet("Summary")

    # Title
    ws_summary['A1'] = "MAYORA INDAH (MYOR IJ) - COMPARABLE COMPANIES ANALYSIS"
    ws_summary['A1'].font = TITLE_FONT
    ws_summary.merge_cells('A1:H1')

    ws_summary['A2'] = "Indonesian FMCG Sector | December 2024 | Currency: IDR"
    ws_summary['A2'].font = SUBHEADER_FONT
    ws_summary.merge_cells('A2:H2')

    # Key metrics box
    ws_summary['A4'] = "INVESTMENT SNAPSHOT"
    ws_summary['A4'].font = HEADER_FONT
    ws_summary['A4'].fill = HEADER_FILL
    ws_summary.merge_cells('A4:D4')

    snapshot_data = [
        ['Current Price', 'IDR 2,120', 'P/E (TTM)', '15.8x'],
        ['52-Week Range', 'IDR 1,820 - 3,010', 'EV/EBITDA', '17.1x'],
        ['Market Cap', 'IDR 47.4T', 'P/B', '2.9x'],
        ['Enterprise Value', 'IDR 68.3T', 'Div Yield', '2.5%'],
    ]

    for i, row in enumerate(snapshot_data, start=5):
        for j, val in enumerate(row):
            cell = ws_summary.cell(row=i, column=j+1, value=val)
            cell.font = DATA_FONT if j % 2 == 1 else SUBHEADER_FONT
            cell.border = THIN_BORDER

    # Valuation summary
    ws_summary['A10'] = "COMPS-IMPLIED VALUATION"
    ws_summary['A10'].font = HEADER_FONT
    ws_summary['A10'].fill = HEADER_FILL
    ws_summary.merge_cells('A10:D10')

    val_data = [
        ['Scenario', 'Implied Price', 'vs Current', 'Probability'],
        ['Bear Case', 'IDR 1,350', '-36%', '25%'],
        ['Base Case', 'IDR 1,700', '-20%', '50%'],
        ['Bull Case', 'IDR 2,100', '-1%', '25%'],
        ['Analyst Consensus', 'IDR 2,747', '+30%', '-'],
    ]

    for i, row in enumerate(val_data, start=11):
        for j, val in enumerate(row):
            cell = ws_summary.cell(row=i, column=j+1, value=val)
            if i == 11:
                cell.font = SUBHEADER_FONT
                cell.fill = SUBHEADER_FILL
            else:
                cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    # ==========================================================================
    # SHEET 2: MARKET DATA
    # ==========================================================================
    ws_market = wb.create_sheet("Market Data")

    df_market = pd.DataFrame(companies)

    # Headers
    ws_market['A1'] = "MARKET DATA SUMMARY"
    ws_market['A1'].font = TITLE_FONT
    ws_market.merge_cells('A1:K1')

    headers = ['Company', 'Ticker', 'Weight', 'Price (IDR)', 'Shares (B)',
               'Mkt Cap (T)', 'Net Debt (T)', 'EV (T)', '52Wk High', '52Wk Low', '52Wk Ret']

    for j, header in enumerate(headers, start=1):
        cell = ws_market.cell(row=3, column=j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # Data
    for i, company in enumerate(companies['Company'], start=4):
        idx = i - 4
        row_data = [
            companies['Company'][idx],
            companies['Ticker'][idx],
            companies['Weight'][idx],
            companies['Price_IDR'][idx],
            companies['Shares_B'][idx],
            companies['Market_Cap_T'][idx],
            companies['Net_Debt_T'][idx],
            companies['EV_T'][idx],
            companies['52Wk_High'][idx],
            companies['52Wk_Low'][idx],
            companies['52Wk_Return'][idx],
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_market.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j >= 4:
                cell.alignment = RIGHT
                if j == 11:  # 52Wk Return
                    cell.number_format = '0.0%'
            if idx == 0:  # Target company
                cell.fill = TARGET_FILL
                cell.font = HIGHLIGHT_FONT

    # ==========================================================================
    # SHEET 3: VALUATION MULTIPLES
    # ==========================================================================
    ws_mult = wb.create_sheet("Valuation Multiples")

    ws_mult['A1'] = "TRADING MULTIPLES COMPARISON"
    ws_mult['A1'].font = TITLE_FONT
    ws_mult.merge_cells('A1:H1')

    mult_headers = ['Company', 'EV/Revenue', 'EV/EBITDA', 'P/E (TTM)', 'P/E (FWD)', 'P/B', 'Div Yield']

    for j, header in enumerate(mult_headers, start=1):
        cell = ws_mult.cell(row=3, column=j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    for i, company in enumerate(multiples['Company'], start=4):
        idx = i - 4
        row_data = [
            multiples['Company'][idx],
            multiples['EV_Revenue'][idx],
            multiples['EV_EBITDA'][idx],
            multiples['PE_TTM'][idx],
            multiples['PE_FWD'][idx] if multiples['PE_FWD'][idx] else 'NM',
            multiples['PB'][idx],
            multiples['Div_Yield'][idx],
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_mult.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j >= 2:
                cell.alignment = RIGHT
                if j == 7:
                    cell.number_format = '0.00%'
                elif isinstance(val, (int, float)):
                    cell.number_format = '0.00x' if j != 7 else '0.00%'
            if idx == 0:
                cell.fill = TARGET_FILL
                cell.font = HIGHLIGHT_FONT

    # Summary stats
    row = 12
    ws_mult.cell(row=row, column=1, value="SUMMARY STATISTICS (Excl. MYOR)").font = SUBHEADER_FONT
    ws_mult.merge_cells('A12:G12')

    # Calculate means and medians
    peer_ev_rev = multiples['EV_Revenue'][1:]
    peer_ev_ebitda = multiples['EV_EBITDA'][1:]
    peer_pe = multiples['PE_TTM'][1:]
    peer_pe_fwd = [x for x in multiples['PE_FWD'][1:] if x]
    peer_pb = multiples['PB'][1:]
    peer_div = multiples['Div_Yield'][1:]

    stats = [
        ['Mean', np.mean(peer_ev_rev), np.mean(peer_ev_ebitda), np.mean(peer_pe),
         np.mean(peer_pe_fwd), np.mean(peer_pb), np.mean(peer_div)],
        ['Median', np.median(peer_ev_rev), np.median(peer_ev_ebitda), np.median(peer_pe),
         np.median(peer_pe_fwd), np.median(peer_pb), np.median(peer_div)],
    ]

    for i, stat_row in enumerate(stats, start=row+1):
        for j, val in enumerate(stat_row, start=1):
            cell = ws_mult.cell(row=i, column=j, value=val)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER
            if j >= 2:
                cell.alignment = RIGHT
                if j == 7:
                    cell.number_format = '0.00%'
                elif isinstance(val, (int, float)):
                    cell.number_format = '0.00x'

    # Premium/Discount
    row = 16
    ws_mult.cell(row=row, column=1, value="MYOR vs Median").font = HIGHLIGHT_FONT
    ws_mult.cell(row=row, column=1).fill = HIGHLIGHT_FILL

    myor_vals = [multiples['EV_Revenue'][0], multiples['EV_EBITDA'][0], multiples['PE_TTM'][0],
                 multiples['PE_FWD'][0], multiples['PB'][0], multiples['Div_Yield'][0]]
    medians = [np.median(peer_ev_rev), np.median(peer_ev_ebitda), np.median(peer_pe),
               np.median(peer_pe_fwd), np.median(peer_pb), np.median(peer_div)]

    for j, (myor, med) in enumerate(zip(myor_vals, medians), start=2):
        if myor and med:
            prem = (myor / med - 1)
            cell = ws_mult.cell(row=row, column=j, value=prem)
            cell.number_format = '+0%;-0%'
            cell.font = HIGHLIGHT_FONT
            cell.fill = HIGHLIGHT_FILL
            cell.border = THIN_BORDER
            cell.alignment = RIGHT

    # ==========================================================================
    # SHEET 4: PROFITABILITY
    # ==========================================================================
    ws_prof = wb.create_sheet("Profitability")

    ws_prof['A1'] = "PROFITABILITY ANALYSIS"
    ws_prof['A1'].font = TITLE_FONT
    ws_prof.merge_cells('A1:G1')

    prof_headers = ['Company', 'Gross Margin', 'EBITDA Margin', 'Net Margin', 'ROE', 'ROIC']

    for j, header in enumerate(prof_headers, start=1):
        cell = ws_prof.cell(row=3, column=j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    for i, company in enumerate(financials['Company'], start=4):
        idx = i - 4
        row_data = [
            financials['Company'][idx],
            financials['Gross_Margin'][idx],
            financials['EBITDA_Margin'][idx],
            financials['Net_Margin'][idx],
            financials['ROE'][idx],
            financials['ROIC'][idx],
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_prof.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j >= 2:
                cell.alignment = RIGHT
                cell.number_format = '0.0%'
            if idx == 0:
                cell.fill = TARGET_FILL
                cell.font = HIGHLIGHT_FONT

    # ==========================================================================
    # SHEET 5: GROWTH
    # ==========================================================================
    ws_growth = wb.create_sheet("Growth")

    ws_growth['A1'] = "GROWTH METRICS"
    ws_growth['A1'].font = TITLE_FONT
    ws_growth.merge_cells('A1:G1')

    growth_headers = ['Company', 'Rev YoY', 'Rev 3Y CAGR', 'EBITDA YoY', 'EPS YoY', 'Fwd Rev', 'Fwd EPS']

    for j, header in enumerate(growth_headers, start=1):
        cell = ws_growth.cell(row=3, column=j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    for i, company in enumerate(growth['Company'], start=4):
        idx = i - 4
        row_data = [
            growth['Company'][idx],
            growth['Rev_Growth_YoY'][idx],
            growth['Rev_CAGR_3Y'][idx],
            growth['EBITDA_Growth_YoY'][idx],
            growth['EPS_Growth_YoY'][idx],
            growth['Fwd_Rev_Growth'][idx],
            growth['Fwd_EPS_Growth'][idx] if growth['Fwd_EPS_Growth'][idx] else 'NM',
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_growth.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j >= 2:
                cell.alignment = RIGHT
                if isinstance(val, (int, float)):
                    cell.number_format = '0.0%'
            if idx == 0:
                cell.fill = TARGET_FILL
                cell.font = HIGHLIGHT_FONT

    # ==========================================================================
    # SHEET 6: IMPLIED VALUATION
    # ==========================================================================
    ws_val = wb.create_sheet("Implied Valuation")

    ws_val['A1'] = "IMPLIED VALUATION ANALYSIS"
    ws_val['A1'].font = TITLE_FONT
    ws_val.merge_cells('A1:F1')

    # MYOR inputs
    ws_val['A3'] = "MYOR INPUTS"
    ws_val['A3'].font = HEADER_FONT
    ws_val['A3'].fill = HEADER_FILL
    ws_val.merge_cells('A3:C3')

    inputs = [
        ['LTM Revenue (T)', 36.07, 'IDR'],
        ['LTM EBITDA (T)', 4.00, 'IDR'],
        ['LTM Net Income (T)', 3.00, 'IDR'],
        ['Net Debt (T)', 20.90, 'IDR'],
        ['Shares Out (B)', 22.36, ''],
        ['Current Price', 2120, 'IDR'],
    ]

    for i, row in enumerate(inputs, start=4):
        for j, val in enumerate(row, start=1):
            cell = ws_val.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # Implied values
    ws_val['A11'] = "IMPLIED VALUE USING PEER MULTIPLES"
    ws_val['A11'].font = HEADER_FONT
    ws_val['A11'].fill = HEADER_FILL
    ws_val.merge_cells('A11:F11')

    imp_headers = ['Multiple', 'MYOR Metric', 'Mean Mult', 'Implied (Mean)', 'Med Mult', 'Implied (Med)']
    for j, h in enumerate(imp_headers, start=1):
        cell = ws_val.cell(row=12, column=j, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    # Calculate implied values
    myor_rev = 36.07
    myor_ebitda = 4.00
    myor_ni = 3.00
    net_debt = 20.90
    shares = 22.36

    mean_ev_rev = np.mean(peer_ev_rev)
    mean_ev_ebitda = np.mean(peer_ev_ebitda)
    mean_pe = np.mean(peer_pe)

    med_ev_rev = np.median(peer_ev_rev)
    med_ev_ebitda = np.median(peer_ev_ebitda)
    med_pe = np.median(peer_pe)

    implied_data = [
        ['EV/Revenue', myor_rev, mean_ev_rev, (myor_rev * mean_ev_rev - net_debt) / shares * 1000,
         med_ev_rev, (myor_rev * med_ev_rev - net_debt) / shares * 1000],
        ['EV/EBITDA', myor_ebitda, mean_ev_ebitda, (myor_ebitda * mean_ev_ebitda - net_debt) / shares * 1000,
         med_ev_ebitda, (myor_ebitda * med_ev_ebitda - net_debt) / shares * 1000],
        ['P/E', myor_ni, mean_pe, myor_ni * mean_pe / shares * 1000,
         med_pe, myor_ni * med_pe / shares * 1000],
    ]

    for i, row in enumerate(implied_data, start=13):
        for j, val in enumerate(row, start=1):
            cell = ws_val.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j in [3, 5]:
                cell.number_format = '0.00x'
            elif j in [4, 6]:
                cell.number_format = '#,##0'

    # Scenario analysis
    ws_val['A18'] = "SCENARIO ANALYSIS"
    ws_val['A18'].font = HEADER_FONT
    ws_val['A18'].fill = HEADER_FILL
    ws_val.merge_cells('A18:E18')

    scen_headers = ['Scenario', 'EV/EBITDA', 'Implied Price', 'Probability', 'Weighted']
    for j, h in enumerate(scen_headers, start=1):
        cell = ws_val.cell(row=19, column=j, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER

    scenarios = [
        ['Bear', 9.0, (4.0 * 9.0 - 20.9) / 22.36 * 1000, 0.25],
        ['Base', 11.0, (4.0 * 11.0 - 20.9) / 22.36 * 1000, 0.50],
        ['Bull', 14.0, (4.0 * 14.0 - 20.9) / 22.36 * 1000, 0.25],
    ]

    for i, row in enumerate(scenarios, start=20):
        for j, val in enumerate(row, start=1):
            cell = ws_val.cell(row=i, column=j, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if j == 2:
                cell.number_format = '0.0x'
            elif j == 3:
                cell.number_format = '#,##0'
            elif j == 4:
                cell.number_format = '0%'
        # Weighted
        cell = ws_val.cell(row=i, column=5, value=row[2] * row[3])
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'

    # Probability weighted
    prob_weighted = sum([s[2] * s[3] for s in scenarios])
    ws_val.cell(row=23, column=1, value="Prob-Weighted Price").font = HIGHLIGHT_FONT
    ws_val.cell(row=23, column=1).fill = HIGHLIGHT_FILL
    ws_val.cell(row=23, column=5, value=prob_weighted).font = HIGHLIGHT_FONT
    ws_val.cell(row=23, column=5).fill = HIGHLIGHT_FILL
    ws_val.cell(row=23, column=5).number_format = '#,##0'

    # ==========================================================================
    # SHEET 7: SENSITIVITY
    # ==========================================================================
    ws_sens = wb.create_sheet("Sensitivity")

    ws_sens['A1'] = "SENSITIVITY ANALYSIS"
    ws_sens['A1'].font = TITLE_FONT
    ws_sens.merge_cells('A1:I1')

    ws_sens['A3'] = "EV/EBITDA SENSITIVITY - IMPLIED SHARE PRICE (IDR)"
    ws_sens['A3'].font = SUBHEADER_FONT
    ws_sens.merge_cells('A3:I3')

    # EBITDA values across top
    ebitda_vals = [3.50, 3.75, 4.00, 4.25, 4.50, 4.75, 5.00]
    ev_ebitda_vals = [8.0, 9.0, 10.0, 11.0, 12.0, 13.0, 14.0]

    ws_sens.cell(row=4, column=1, value="EV/EBITDA").font = HEADER_FONT
    ws_sens.cell(row=4, column=1).fill = HEADER_FILL

    for j, ebitda in enumerate(ebitda_vals, start=2):
        cell = ws_sens.cell(row=4, column=j, value=ebitda)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.number_format = '0.00'

    for i, mult in enumerate(ev_ebitda_vals, start=5):
        cell = ws_sens.cell(row=i, column=1, value=mult)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.number_format = '0.0x'

        for j, ebitda in enumerate(ebitda_vals, start=2):
            implied = (ebitda * mult - 20.9) / 22.36 * 1000
            cell = ws_sens.cell(row=i, column=j, value=implied)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.number_format = '#,##0'
            cell.alignment = CENTER

            # Highlight current (4.0 EBITDA, ~10-11x)
            if ebitda == 4.00 and mult in [10.0, 11.0]:
                cell.fill = HIGHLIGHT_FILL

    # ==========================================================================
    # ADJUST COLUMN WIDTHS
    # ==========================================================================
    for ws in wb.worksheets:
        for col_idx in range(1, 15):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15

    # ==========================================================================
    # SAVE
    # ==========================================================================
    output_path = '/Users/nathanielluu/natan-equity-research/natan-equity-research/MYOR_Comps_Model.xlsx'
    wb.save(output_path)
    print("Excel model saved to: " + output_path)
    return output_path

if __name__ == "__main__":
    try:
        create_excel_model()
        print("\nModel created successfully!")
        print("\nSheets included:")
        print("  1. Summary - Investment snapshot and valuation summary")
        print("  2. Market Data - Stock prices, market caps, enterprise values")
        print("  3. Valuation Multiples - EV/EBITDA, P/E, P/B with peer stats")
        print("  4. Profitability - Margins, ROE, ROIC analysis")
        print("  5. Growth - Revenue and earnings growth metrics")
        print("  6. Implied Valuation - Price targets using peer multiples")
        print("  7. Sensitivity - EV/EBITDA sensitivity table")
    except ImportError as e:
        print("Missing required library: " + str(e))
        print("\nInstall with: pip install openpyxl pandas numpy")
